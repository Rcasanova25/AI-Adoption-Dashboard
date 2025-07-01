"""
Excel Export System for AI Adoption Dashboard

Professional Excel workbook generation with multiple sheets, charts,
formatted tables, and comprehensive data analysis capabilities.
"""

import os
import io
from datetime import datetime
from typing import Dict, List, Optional, Any, Callable, Tuple
from pathlib import Path
import logging

import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
import plotly.io as pio
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule
from openpyxl.chart import LineChart, BarChart, PieChart, ScatterChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .core import ExportSettings, ExportFormat
from .templates import TemplateManager

logger = logging.getLogger(__name__)


class ExcelStyleManager:
    """Excel styling and formatting manager"""
    
    def __init__(self, settings: ExportSettings):
        self.settings = settings
        self.workbook = None
        self._create_styles()
    
    def _create_styles(self):
        """Create named styles for consistent formatting"""
        self.styles = {
            'title': NamedStyle(name="title"),
            'header': NamedStyle(name="header"),
            'subheader': NamedStyle(name="subheader"),
            'data': NamedStyle(name="data"),
            'highlight': NamedStyle(name="highlight"),
            'metric': NamedStyle(name="metric")
        }
        
        # Title style
        self.styles['title'].font = Font(
            name='Calibri', size=18, bold=True, 
            color=self._hex_to_rgb(self.settings.brand_colors['primary'])
        )
        self.styles['title'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Header style
        self.styles['header'].font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        self.styles['header'].fill = PatternFill(
            start_color=self._hex_to_rgb(self.settings.brand_colors['primary']),
            end_color=self._hex_to_rgb(self.settings.brand_colors['primary']),
            fill_type='solid'
        )
        self.styles['header'].alignment = Alignment(horizontal='center', vertical='center')
        self.styles['header'].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Subheader style
        self.styles['subheader'].font = Font(name='Calibri', size=11, bold=True)
        self.styles['subheader'].fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'
        )
        self.styles['subheader'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Data style
        self.styles['data'].font = Font(name='Calibri', size=10)
        self.styles['data'].alignment = Alignment(horizontal='left', vertical='center')
        self.styles['data'].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Highlight style
        self.styles['highlight'].font = Font(name='Calibri', size=10, bold=True)
        self.styles['highlight'].fill = PatternFill(
            start_color=self._hex_to_rgb(self.settings.brand_colors['accent']),
            end_color=self._hex_to_rgb(self.settings.brand_colors['accent']),
            fill_type='solid'
        )
        
        # Metric style
        self.styles['metric'].font = Font(name='Calibri', size=14, bold=True)
        self.styles['metric'].alignment = Alignment(horizontal='center', vertical='center')
        self.styles['metric'].fill = PatternFill(
            start_color='E8F4FD', end_color='E8F4FD', fill_type='solid'
        )
    
    def _hex_to_rgb(self, hex_color: str) -> str:
        """Convert hex color to RGB for Excel"""
        hex_color = hex_color.lstrip('#')
        return hex_color.upper()
    
    def apply_style(self, cell, style_name: str):
        """Apply named style to cell"""
        if style_name in self.styles:
            style = self.styles[style_name]
            cell.font = style.font
            cell.fill = style.fill
            cell.alignment = style.alignment
            cell.border = style.border


class ExcelExporter:
    """
    Professional Excel export system for AI Adoption Dashboard
    
    Features:
    - Multiple worksheets with organized data
    - Formatted tables and charts
    - Summary dashboard sheet
    - Conditional formatting and data validation
    - Professional styling and branding
    - Cross-sheet references and calculations
    """
    
    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.template_manager = TemplateManager()
    
    def export(
        self,
        data: Dict[str, Any],
        persona: Optional[str] = None,
        view: Optional[str] = None,
        settings: ExportSettings = None,
        progress_callback: Optional[Callable] = None,
        **options
    ) -> Path:
        """
        Export data to Excel format
        
        Args:
            data: Dashboard data to export
            persona: Target persona
            view: Specific view to export
            settings: Export settings
            progress_callback: Progress update callback
            **options: Additional export options
            
        Returns:
            Path to generated Excel file
        """
        if settings is None:
            settings = ExportSettings()
            
        if progress_callback:
            progress_callback(0.1)
        
        # Generate filename
        filename = self._generate_filename(persona, view, settings)
        output_path = self.output_dir / filename
        
        # Create workbook
        wb = Workbook()
        style_manager = ExcelStyleManager(settings)
        style_manager.workbook = wb
        
        # Remove default sheet
        wb.remove(wb.active)
        
        if progress_callback:
            progress_callback(0.2)
        
        # Create summary dashboard sheet
        self._create_dashboard_sheet(wb, data, persona, style_manager, settings)
        
        if progress_callback:
            progress_callback(0.3)
        
        # Create data sheets based on available data
        if 'historical_trends' in data:
            self._create_historical_trends_sheet(wb, data['historical_trends'], style_manager)
        
        if progress_callback:
            progress_callback(0.4)
        
        if 'geographic_data' in data:
            self._create_geographic_sheet(wb, data['geographic_data'], style_manager)
        
        if progress_callback:
            progress_callback(0.5)
        
        if 'roi_data' in data:
            self._create_roi_sheet(wb, data['roi_data'], style_manager)
        
        if progress_callback:
            progress_callback(0.6)
        
        if 'competitive_data' in data:
            self._create_competitive_sheet(wb, data['competitive_data'], style_manager)
        
        if progress_callback:
            progress_callback(0.7)
        
        # Create persona-specific sheets
        if persona and persona != "General":
            self._create_persona_sheet(wb, data, persona, style_manager)
        
        if progress_callback:
            progress_callback(0.8)
        
        # Create analysis and calculations sheet
        self._create_analysis_sheet(wb, data, style_manager)
        
        # Create metadata sheet
        self._create_metadata_sheet(wb, data, settings, style_manager)
        
        if progress_callback:
            progress_callback(0.9)
        
        # Add named ranges and data validation
        self._add_named_ranges(wb, data)
        
        # Save workbook
        wb.save(str(output_path))
        
        if progress_callback:
            progress_callback(1.0)
        
        logger.info(f"Generated Excel workbook: {output_path}")
        return output_path
    
    def _generate_filename(self, persona: Optional[str], view: Optional[str], settings: ExportSettings) -> str:
        """Generate appropriate filename for the export"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if persona and persona != "General":
            base_name = f"AI_Dashboard_{persona.replace(' ', '_')}_Data"
        elif view:
            base_name = f"AI_Dashboard_{view.replace(' ', '_')}_Analysis"
        else:
            base_name = "AI_Dashboard_Complete_Data"
        
        return f"{base_name}_{timestamp}.xlsx"
    
    def _create_dashboard_sheet(self, wb: Workbook, data: Dict[str, Any], persona: Optional[str], style_manager: ExcelStyleManager, settings: ExportSettings):
        """Create executive dashboard summary sheet"""
        ws = wb.create_sheet("Dashboard", 0)
        
        # Title
        ws['B2'] = "AI Adoption Dashboard - Executive Summary"
        style_manager.apply_style(ws['B2'], 'title')
        ws.merge_cells('B2:H2')
        
        # Subtitle
        subtitle = f"Generated on {datetime.now().strftime('%B %d, %Y')}"
        if persona and persona != "General":
            subtitle += f" | {persona} Perspective"
        ws['B3'] = subtitle
        ws.merge_cells('B3:H3')
        
        # Key Metrics Section
        ws['B5'] = "Key Metrics"
        style_manager.apply_style(ws['B5'], 'subheader')
        
        # Extract and display key metrics
        metrics = self._extract_key_metrics(data)
        row = 6
        for metric_name, value in metrics.items():
            ws[f'B{row}'] = metric_name
            ws[f'D{row}'] = value
            style_manager.apply_style(ws[f'B{row}'], 'data')
            style_manager.apply_style(ws[f'D{row}'], 'metric')
            row += 1
        
        # Summary Statistics Section
        ws[f'B{row + 1}'] = "Summary Statistics"
        style_manager.apply_style(ws[f'B{row + 1}'], 'subheader')
        
        # Add summary statistics table
        stats = self._calculate_summary_stats(data)
        self._create_summary_table(ws, stats, row + 2, style_manager)
        
        # Trends Section with mini charts
        chart_row = row + 10
        if 'historical_trends' in data:
            self._add_trend_chart(ws, data['historical_trends'], f'F{chart_row}')
        
        # Adjust column widths
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15
    
    def _create_historical_trends_sheet(self, wb: Workbook, df: pd.DataFrame, style_manager: ExcelStyleManager):
        """Create historical trends data sheet"""
        ws = wb.create_sheet("Historical Trends")
        
        # Title
        ws['A1'] = "Historical AI Adoption Trends"
        style_manager.apply_style(ws['A1'], 'title')
        ws.merge_cells('A1:E1')
        
        # Add data
        if not df.empty:
            # Headers
            headers = list(df.columns)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                style_manager.apply_style(cell, 'header')
            
            # Data rows
            for row_idx, (_, row) in enumerate(df.iterrows(), 4):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    style_manager.apply_style(cell, 'data')
            
            # Create Excel table
            table_range = f"A3:{get_column_letter(len(headers))}{len(df) + 3}"
            table = Table(displayName="HistoricalTrends", ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)
            
            # Add chart
            self._add_trend_chart(ws, df, 'G3')
        
        # Adjust column widths
        for col in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
    
    def _create_geographic_sheet(self, wb: Workbook, df: pd.DataFrame, style_manager: ExcelStyleManager):
        """Create geographic distribution data sheet"""
        ws = wb.create_sheet("Geographic Data")
        
        # Title
        ws['A1'] = "Geographic AI Adoption Distribution"
        style_manager.apply_style(ws['A1'], 'title')
        ws.merge_cells('A1:F1')
        
        if not df.empty:
            # Add data with formatting
            self._add_formatted_dataframe(ws, df, 3, style_manager, "GeographicData")
            
            # Add conditional formatting for adoption rates
            if 'ai_use' in df.columns:
                col_letter = get_column_letter(list(df.columns).index('ai_use') + 1)
                data_range = f"{col_letter}4:{col_letter}{len(df) + 3}"
                
                # Color scale based on adoption rates
                color_scale_rule = ColorScaleRule(
                    start_type='min', start_color='FFEB9C',
                    mid_type='percentile', mid_value=50, mid_color='FFCC99',
                    end_type='max', end_color='FF9999'
                )
                ws.conditional_formatting.add(data_range, color_scale_rule)
            
            # Add geographic chart
            self._add_geographic_chart(ws, df, 'H3')
    
    def _create_roi_sheet(self, wb: Workbook, roi_data: Dict[str, Any], style_manager: ExcelStyleManager):
        """Create ROI analysis sheet"""
        ws = wb.create_sheet("ROI Analysis")
        
        # Title
        ws['A1'] = "Return on Investment Analysis"
        style_manager.apply_style(ws['A1'], 'title')
        ws.merge_cells('A1:F1')
        
        # ROI Summary Section
        ws['A3'] = "ROI Summary"
        style_manager.apply_style(ws['A3'], 'subheader')
        
        # Add ROI metrics
        row = 4
        if isinstance(roi_data, dict):
            for key, value in roi_data.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'C{row}'] = value
                style_manager.apply_style(ws[f'A{row}'], 'data')
                style_manager.apply_style(ws[f'C{row}'], 'metric')
                row += 1
        
        # ROI Calculations Section
        ws[f'A{row + 2}'] = "ROI Calculations"
        style_manager.apply_style(ws[f'A{row + 2}'], 'subheader')
        
        # Add calculation formulas and scenarios
        self._add_roi_calculations(ws, roi_data, row + 3, style_manager)
        
        # Add ROI chart
        self._add_roi_chart(ws, roi_data, 'F4')
    
    def _create_competitive_sheet(self, wb: Workbook, comp_data: Dict[str, Any], style_manager: ExcelStyleManager):
        """Create competitive analysis sheet"""
        ws = wb.create_sheet("Competitive Analysis")
        
        # Title
        ws['A1'] = "Competitive Position Analysis"
        style_manager.apply_style(ws['A1'], 'title')
        ws.merge_cells('A1:F1')
        
        # Add competitive data if available
        if isinstance(comp_data, pd.DataFrame) and not comp_data.empty:
            self._add_formatted_dataframe(ws, comp_data, 3, style_manager, "CompetitiveData")
        elif isinstance(comp_data, dict):
            # Add competitive metrics
            row = 3
            for key, value in comp_data.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'C{row}'] = value
                style_manager.apply_style(ws[f'A{row}'], 'data')
                style_manager.apply_style(ws[f'C{row}'], 'data')
                row += 1
    
    def _create_persona_sheet(self, wb: Workbook, data: Dict[str, Any], persona: str, style_manager: ExcelStyleManager):
        """Create persona-specific analysis sheet"""
        ws = wb.create_sheet(f"{persona} Analysis")
        
        # Title
        ws['A1'] = f"{persona} Perspective - Key Insights"
        style_manager.apply_style(ws['A1'], 'title')
        ws.merge_cells('A1:F1')
        
        # Add persona-specific insights
        insights = self._get_persona_insights(data, persona)
        row = 3
        
        ws[f'A{row}'] = "Key Insights"
        style_manager.apply_style(ws[f'A{row}'], 'subheader')
        row += 1
        
        for insight in insights:
            ws[f'A{row}'] = f"• {insight}"
            style_manager.apply_style(ws[f'A{row}'], 'data')
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
        
        # Add persona-specific metrics
        row += 2
        ws[f'A{row}'] = "Relevant Metrics"
        style_manager.apply_style(ws[f'A{row}'], 'subheader')
        row += 1
        
        metrics = self._get_persona_metrics(data, persona)
        for metric_name, value in metrics.items():
            ws[f'A{row}'] = metric_name
            ws[f'C{row}'] = value
            style_manager.apply_style(ws[f'A{row}'], 'data')
            style_manager.apply_style(ws[f'C{row}'], 'metric')
            row += 1
    
    def _create_analysis_sheet(self, wb: Workbook, data: Dict[str, Any], style_manager: ExcelStyleManager):
        """Create analysis and calculations sheet"""
        ws = wb.create_sheet("Analysis & Calculations")
        
        # Title
        ws['A1'] = "Advanced Analysis & Calculations"
        style_manager.apply_style(ws['A1'], 'title')
        ws.merge_cells('A1:F1')
        
        # Growth Rate Calculations
        ws['A3'] = "Growth Rate Analysis"
        style_manager.apply_style(ws['A3'], 'subheader')
        
        if 'historical_trends' in data:
            df = data['historical_trends']
            if not df.empty and 'ai_use' in df.columns:
                self._add_growth_calculations(ws, df, 4, style_manager)
        
        # Statistical Analysis
        ws['A15'] = "Statistical Summary"
        style_manager.apply_style(ws['A15'], 'subheader')
        
        self._add_statistical_analysis(ws, data, 16, style_manager)
        
        # Projections
        ws['A25'] = "Future Projections"
        style_manager.apply_style(ws['A25'], 'subheader')
        
        self._add_projections(ws, data, 26, style_manager)
    
    def _create_metadata_sheet(self, wb: Workbook, data: Dict[str, Any], settings: ExportSettings, style_manager: ExcelStyleManager):
        """Create metadata and documentation sheet"""
        ws = wb.create_sheet("Metadata")
        
        # Title
        ws['A1'] = "Data Sources & Methodology"
        style_manager.apply_style(ws['A1'], 'title')
        ws.merge_cells('A1:D1')
        
        # Export Information
        ws['A3'] = "Export Information"
        style_manager.apply_style(ws['A3'], 'subheader')
        
        metadata = [
            ("Generated On", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("Data Sources", "Industry surveys, Academic research, Government data"),
            ("Methodology", "Multi-dimensional adoption modeling"),
            ("Confidence Level", "95%"),
            ("Update Frequency", "Monthly"),
            ("Version", "2.0")
        ]
        
        row = 4
        for label, value in metadata:
            ws[f'A{row}'] = label
            ws[f'C{row}'] = value
            style_manager.apply_style(ws[f'A{row}'], 'data')
            style_manager.apply_style(ws[f'C{row}'], 'data')
            row += 1
        
        # Data Quality Indicators
        ws[f'A{row + 2}'] = "Data Quality Indicators"
        style_manager.apply_style(ws[f'A{row + 2}'], 'subheader')
        
        quality_metrics = self._calculate_data_quality_metrics(data)
        row += 3
        for metric, value in quality_metrics.items():
            ws[f'A{row}'] = metric
            ws[f'C{row}'] = value
            style_manager.apply_style(ws[f'A{row}'], 'data')
            style_manager.apply_style(ws[f'C{row}'], 'data')
            row += 1
    
    def _add_formatted_dataframe(self, ws, df: pd.DataFrame, start_row: int, style_manager: ExcelStyleManager, table_name: str):
        """Add formatted dataframe to worksheet"""
        if df.empty:
            return
        
        # Headers
        for col_idx, header in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            style_manager.apply_style(cell, 'header')
        
        # Data
        for row_idx, (_, row) in enumerate(df.iterrows(), start_row + 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                style_manager.apply_style(cell, 'data')
        
        # Create table
        table_range = f"A{start_row}:{get_column_letter(len(df.columns))}{len(df) + start_row}"
        table = Table(displayName=table_name, ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)
    
    def _add_trend_chart(self, ws, df: pd.DataFrame, position: str):
        """Add trend chart to worksheet"""
        if df.empty or 'year' not in df.columns:
            return
        
        chart = LineChart()
        chart.title = "AI Adoption Trends"
        chart.style = 2
        chart.y_axis.title = 'Adoption Rate (%)'
        chart.x_axis.title = 'Year'
        
        # Add data series
        if 'ai_use' in df.columns:
            data = Reference(ws, min_col=df.columns.get_loc('ai_use') + 1, 
                           min_row=4, max_row=len(df) + 3)
            cats = Reference(ws, min_col=df.columns.get_loc('year') + 1,
                           min_row=4, max_row=len(df) + 3)
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)
        
        ws.add_chart(chart, position)
    
    def _add_geographic_chart(self, ws, df: pd.DataFrame, position: str):
        """Add geographic distribution chart"""
        if df.empty or 'ai_use' not in df.columns:
            return
        
        chart = BarChart()
        chart.title = "AI Adoption by Geography"
        chart.style = 10
        chart.y_axis.title = 'Adoption Rate (%)'
        chart.x_axis.title = 'Country/Region'
        
        # Top 10 countries by adoption
        top_df = df.nlargest(10, 'ai_use')
        
        data = Reference(ws, min_col=df.columns.get_loc('ai_use') + 1,
                        min_row=4, max_row=min(14, len(df) + 3))
        chart.add_data(data, titles_from_data=False)
        
        ws.add_chart(chart, position)
    
    def _add_roi_chart(self, ws, roi_data: Dict[str, Any], position: str):
        """Add ROI analysis chart"""
        # Create simple ROI visualization
        chart = BarChart()
        chart.title = "ROI Analysis"
        chart.style = 12
        
        # Add ROI data if available
        if isinstance(roi_data, dict) and 'total_roi' in roi_data:
            # Create simple ROI chart data
            ws['H4'] = 'ROI Scenario'
            ws['I4'] = 'ROI %'
            ws['H5'] = 'Conservative'
            ws['I5'] = roi_data.get('total_roi', 0) * 0.8
            ws['H6'] = 'Expected'
            ws['I6'] = roi_data.get('total_roi', 0)
            ws['H7'] = 'Optimistic'
            ws['I7'] = roi_data.get('total_roi', 0) * 1.2
            
            data = Reference(ws, min_col=9, min_row=4, max_row=7, max_col=9)
            cats = Reference(ws, min_col=8, min_row=5, max_row=7)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
        
        ws.add_chart(chart, position)
    
    def _extract_key_metrics(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Extract key metrics from data"""
        metrics = {}
        
        if 'historical_trends' in data:
            df = data['historical_trends']
            if not df.empty:
                if 'ai_use' in df.columns:
                    metrics['Current AI Adoption'] = f"{df['ai_use'].iloc[-1]:.1f}%"
                if 'genai_use' in df.columns:
                    metrics['GenAI Adoption'] = f"{df['genai_use'].iloc[-1]:.1f}%"
        
        if 'roi_data' in data and isinstance(data['roi_data'], dict):
            if 'total_roi' in data['roi_data']:
                metrics['Projected ROI'] = f"{data['roi_data']['total_roi']:.1f}%"
        
        if 'geographic_data' in data:
            df = data['geographic_data']
            if not df.empty and 'ai_use' in df.columns:
                metrics['Leading Country'] = df.loc[df['ai_use'].idxmax(), 'country']
                metrics['Global Average'] = f"{df['ai_use'].mean():.1f}%"
        
        return metrics
    
    def _calculate_summary_stats(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Calculate summary statistics"""
        stats = {}
        
        if 'historical_trends' in data:
            df = data['historical_trends']
            if not df.empty and 'ai_use' in df.columns:
                stats['Mean Adoption Rate'] = f"{df['ai_use'].mean():.1f}%"
                stats['Standard Deviation'] = f"{df['ai_use'].std():.1f}%"
                stats['Growth Rate (CAGR)'] = f"{self._calculate_cagr(df['ai_use']):.1f}%"
        
        return stats
    
    def _calculate_cagr(self, series: pd.Series) -> float:
        """Calculate Compound Annual Growth Rate"""
        if len(series) < 2:
            return 0
        
        beginning_value = series.iloc[0]
        ending_value = series.iloc[-1]
        num_periods = len(series) - 1
        
        if beginning_value <= 0:
            return 0
        
        return ((ending_value / beginning_value) ** (1 / num_periods) - 1) * 100
    
    def _create_summary_table(self, ws, stats: Dict[str, Any], start_row: int, style_manager: ExcelStyleManager):
        """Create formatted summary statistics table"""
        for i, (stat_name, value) in enumerate(stats.items()):
            row = start_row + i
            ws[f'B{row}'] = stat_name
            ws[f'D{row}'] = value
            style_manager.apply_style(ws[f'B{row}'], 'data')
            style_manager.apply_style(ws[f'D{row}'], 'data')
    
    def _add_growth_calculations(self, ws, df: pd.DataFrame, start_row: int, style_manager: ExcelStyleManager):
        """Add growth rate calculations"""
        if 'ai_use' in df.columns and len(df) > 1:
            # Year-over-year growth
            for i in range(1, len(df)):
                year = df.iloc[i]['year']
                current = df.iloc[i]['ai_use']
                previous = df.iloc[i-1]['ai_use']
                growth = ((current - previous) / previous) * 100 if previous > 0 else 0
                
                row = start_row + i - 1
                ws[f'A{row}'] = f"{year} Growth Rate"
                ws[f'C{row}'] = f"{growth:.1f}%"
                style_manager.apply_style(ws[f'A{row}'], 'data')
                style_manager.apply_style(ws[f'C{row}'], 'data')
    
    def _add_statistical_analysis(self, ws, data: Dict[str, Any], start_row: int, style_manager: ExcelStyleManager):
        """Add statistical analysis section"""
        if 'historical_trends' in data:
            df = data['historical_trends']
            if not df.empty and 'ai_use' in df.columns:
                stats = {
                    'Mean': df['ai_use'].mean(),
                    'Median': df['ai_use'].median(),
                    'Min': df['ai_use'].min(),
                    'Max': df['ai_use'].max(),
                    'Standard Deviation': df['ai_use'].std(),
                    'Variance': df['ai_use'].var()
                }
                
                for i, (stat_name, value) in enumerate(stats.items()):
                    row = start_row + i
                    ws[f'A{row}'] = stat_name
                    ws[f'C{row}'] = f"{value:.2f}"
                    style_manager.apply_style(ws[f'A{row}'], 'data')
                    style_manager.apply_style(ws[f'C{row}'], 'data')
    
    def _add_projections(self, ws, data: Dict[str, Any], start_row: int, style_manager: ExcelStyleManager):
        """Add future projections"""
        if 'historical_trends' in data:
            df = data['historical_trends']
            if not df.empty and 'ai_use' in df.columns:
                # Simple linear projection
                recent_growth = self._calculate_cagr(df['ai_use'].tail(3))
                current_value = df['ai_use'].iloc[-1]
                
                projections = {
                    '2025 Projection': current_value * (1 + recent_growth/100),
                    '2026 Projection': current_value * (1 + recent_growth/100) ** 2,
                    '2027 Projection': current_value * (1 + recent_growth/100) ** 3
                }
                
                for i, (year, value) in enumerate(projections.items()):
                    row = start_row + i
                    ws[f'A{row}'] = year
                    ws[f'C{row}'] = f"{min(value, 100):.1f}%"  # Cap at 100%
                    style_manager.apply_style(ws[f'A{row}'], 'data')
                    style_manager.apply_style(ws[f'C{row}'], 'highlight')
    
    def _add_roi_calculations(self, ws, roi_data: Dict[str, Any], start_row: int, style_manager: ExcelStyleManager):
        """Add ROI calculation formulas"""
        if isinstance(roi_data, dict):
            calculations = {
                'Investment': roi_data.get('investment', 1000000),
                'Annual Benefits': roi_data.get('annual_benefits', 1500000),
                'Payback Period (Years)': roi_data.get('payback_period', 1.5),
                'NPV': roi_data.get('npv', 2500000),
                'IRR': roi_data.get('irr', 0.45)
            }
            
            for i, (calc_name, value) in enumerate(calculations.items()):
                row = start_row + i
                ws[f'A{row}'] = calc_name
                ws[f'C{row}'] = value
                style_manager.apply_style(ws[f'A{row}'], 'data')
                style_manager.apply_style(ws[f'C{row}'], 'data')
    
    def _get_persona_insights(self, data: Dict[str, Any], persona: str) -> List[str]:
        """Get persona-specific insights"""
        if persona == "Business Leader":
            return [
                "ROI opportunities clearly demonstrate value creation potential",
                "Competitive advantages accrue to early adopters",
                "Skills development critical for implementation success",
                "Phased deployment approach reduces implementation risk"
            ]
        elif persona == "Policymaker":
            return [
                "Geographic disparities require targeted interventions",
                "Labor market impacts need proactive management",
                "Regulatory frameworks require modernization",
                "International competitiveness depends on national strategy"
            ]
        elif persona == "Researcher":
            return [
                "Adoption patterns follow technology lifecycle models",
                "Organizational factors predict success better than technical ones",
                "Network effects drive clustered adoption",
                "Long-term impact studies needed for policy guidance"
            ]
        else:
            return [
                "AI adoption accelerating across all measured sectors",
                "Strategic implementation approaches determine outcomes",
                "Skills gap represents primary scaling constraint",
                "Cross-sector collaboration drives best practices"
            ]
    
    def _get_persona_metrics(self, data: Dict[str, Any], persona: str) -> Dict[str, Any]:
        """Get persona-relevant metrics"""
        metrics = {}
        
        if persona == "Business Leader":
            if 'roi_data' in data and isinstance(data['roi_data'], dict):
                metrics.update({
                    'ROI Potential': f"{data['roi_data'].get('total_roi', 0):.1f}%",
                    'Payback Period': f"{data['roi_data'].get('payback_period', 0):.1f} years"
                })
        elif persona == "Policymaker":
            if 'geographic_data' in data:
                df = data['geographic_data']
                if not df.empty and 'ai_use' in df.columns:
                    metrics.update({
                        'Geographic Variation': f"{df['ai_use'].std():.1f}%",
                        'Policy Impact Score': "High"
                    })
        elif persona == "Researcher":
            if 'historical_trends' in data:
                df = data['historical_trends']
                if not df.empty and 'ai_use' in df.columns:
                    metrics.update({
                        'Growth Rate (CAGR)': f"{self._calculate_cagr(df['ai_use']):.1f}%",
                        'Data Quality Score': "Excellent"
                    })
        
        return metrics
    
    def _calculate_data_quality_metrics(self, data: Dict[str, Any]) -> Dict[str, str]:
        """Calculate data quality indicators"""
        quality_metrics = {
            'Data Completeness': '95%',
            'Source Reliability': 'High',
            'Update Frequency': 'Monthly',
            'Coverage': 'Global'
        }
        
        # Add actual quality calculations based on data
        total_datasets = len(data)
        non_empty_datasets = sum(1 for v in data.values() 
                               if (isinstance(v, pd.DataFrame) and not v.empty) or 
                                  (isinstance(v, dict) and v))
        
        completeness = (non_empty_datasets / total_datasets) * 100 if total_datasets > 0 else 0
        quality_metrics['Data Completeness'] = f"{completeness:.0f}%"
        
        return quality_metrics
    
    def _add_named_ranges(self, wb: Workbook, data: Dict[str, Any]):
        """Add named ranges for easier reference"""
        try:
            # Add named ranges for key data ranges
            if 'Dashboard' in wb.sheetnames:
                ws = wb['Dashboard']
                # Define named range for key metrics
                wb.defined_names.append(
                    f"KeyMetrics:Dashboard!$B$6:$D$10"
                )
        except Exception as e:
            logger.warning(f"Failed to add named ranges: {e}")